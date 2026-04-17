"""
엑셀 주문서 파서 모듈

다양한 형식의 엑셀 주문서에서 사이즈 목록과 수량을 자동으로 추출한다.
승화전사 유니폼 업계에서 사용되는 주문서 형식은 천차만별이므로,
셀을 전체 스캔하여 사이즈 키워드를 찾고 인접 셀에서 수량을 추출하는 방식을 쓴다.

지원 형식:
  - 가로형: 사이즈가 한 행에 나열 (S | M | L | XL), 아래 행에 수량
  - 세로형: 사이즈가 한 열에 나열 (S 아래 M 아래 L), 오른쪽 열에 수량
  - 표형: 헤더 행에 사이즈, 데이터 행에 수량 (가로형의 변형)
"""

import re
from typing import Optional
import openpyxl

# openpyxl Custom Properties 버그 워크어라운드
# 일부 엑셀 파일에서 Custom Document Properties의 name이 None이면 오류 발생
# (StringProperty.name should be str but value is NoneType)
try:
    from openpyxl.packaging.custom import StringProperty
    _orig_str_init = StringProperty.__init__
    def _patched_str_init(self, *args, **kwargs):
        if 'name' in kwargs and kwargs['name'] is None:
            kwargs['name'] = ''
        _orig_str_init(self, *args, **kwargs)
    StringProperty.__init__ = _patched_str_init
except Exception:
    pass


# 사이즈 키워드 목록 (긴 것부터 매칭해야 "XL" 이 "5XL" 보다 먼저 매칭되는 실수를 방지)
SIZE_KEYWORDS = [
    "5XS", "4XS", "3XS", "2XS", "XS",
    "S", "M",
    "L",
    "XL", "2XL", "3XL", "4XL", "5XL",
]

# 정규식: 셀 텍스트에서 사이즈를 정확하게 추출하기 위한 패턴
# "2XL", "2 XL", "2-XL", "L사이즈", "L 사이즈" 등을 처리
# 긴 키워드부터 먼저 매칭해야 "5XL"이 "XL"로 잘못 매칭되지 않음
_SIZE_PATTERN = re.compile(
    r"^[\s]*("
    + "|".join(
        # 숫자+XS/XL 사이즈는 숫자와 문자 사이에 공백/하이픈 허용
        re.escape(s).replace(r"\d", r"\d")
        for s in sorted(SIZE_KEYWORDS, key=len, reverse=True)
    ).replace("5XS", r"5[\s\-]?XS").replace("4XS", r"4[\s\-]?XS").replace("3XS", r"3[\s\-]?XS").replace("2XS", r"2[\s\-]?XS")
     .replace("5XL", r"5[\s\-]?XL").replace("4XL", r"4[\s\-]?XL").replace("3XL", r"3[\s\-]?XL").replace("2XL", r"2[\s\-]?XL")
    + r")[\s]?(?:사이즈)?[\s]*$",
    re.IGNORECASE,
)


def _normalize_size(text: str) -> Optional[str]:
    """
    셀 텍스트를 정규화된 사이즈 키워드로 변환한다.
    매칭되지 않으면 None을 반환한다.

    예: "2 XL" -> "2XL", "l사이즈" -> "L", "M" -> "M"
    """
    if text is None:
        return None
    text = str(text).strip()
    if not text:
        return None

    m = _SIZE_PATTERN.match(text)
    if m:
        # 매칭된 그룹에서 공백/하이픈을 제거하고 대문자로 통일
        raw = m.group(1).upper().replace(" ", "").replace("-", "")
        # SIZE_KEYWORDS에 있는 정확한 형태로 매핑
        for kw in SIZE_KEYWORDS:
            if raw == kw:
                return kw
        # 숫자+XS/XL 패턴 재확인 (위에서 못 잡은 경우)
        return raw if raw in SIZE_KEYWORDS else None
    return None


def _extract_quantity(cell_value) -> int:
    """
    셀 값에서 수량(정수)을 추출한다.
    숫자가 아니거나 없으면 0을 반환한다.
    """
    if cell_value is None:
        return 0
    if isinstance(cell_value, (int, float)):
        val = int(cell_value)
        # 음수나 비정상적으로 큰 값은 무시
        return val if 0 < val < 100000 else 0
    # 문자열에서 숫자 추출 시도
    text = str(cell_value).strip()
    # "12장", "12개", "12 EA" 같은 패턴에서 숫자만 추출
    num_match = re.match(r"^(\d+)", text)
    if num_match:
        val = int(num_match.group(1))
        return val if 0 < val < 100000 else 0
    return 0


def parse_order_excel(excel_path: str) -> dict:
    """
    주문서 엑셀 파일에서 사이즈 목록과 수량을 추출한다.

    전략:
    1. 전체 셀을 스캔하여 사이즈 키워드가 있는 셀 위치를 모두 수집
    2. 같은 행에 2개 이상 사이즈가 있으면 "가로형", 같은 열이면 "세로형"
    3. 사이즈 셀의 인접 셀(아래 또는 오른쪽)에서 수량 추출
    4. 중복 사이즈는 수량을 합산
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        return {
            "success": False,
            "error": f"엑셀 파일을 열 수 없습니다: {str(e)}",
            "sizes": [],
            "total_quantity": 0,
            "source_sheet": "",
            "detected_format": "unknown",
        }

    best_result = None  # 가장 많은 사이즈를 찾은 시트의 결과를 사용

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 1단계: 모든 셀을 스캔하여 사이즈 키워드 위치 수집
        # size_cells: [(row, col, normalized_size), ...]
        size_cells = []
        # 시트의 모든 행을 순회 (read_only 모드에서는 iter_rows 사용)
        rows_data = []
        for row in ws.iter_rows():
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            rows_data.append(row_values)

        for r_idx, row_vals in enumerate(rows_data):
            for c_idx, val in enumerate(row_vals):
                normalized = _normalize_size(val)
                if normalized:
                    size_cells.append((r_idx, c_idx, normalized))

        if not size_cells:
            continue  # 이 시트에는 사이즈가 없음 → 다음 시트

        # 2단계: 가로형 vs 세로형 판별
        # 같은 행에 있는 사이즈 수 vs 같은 열에 있는 사이즈 수로 판단
        row_counts = {}  # row -> count
        col_counts = {}  # col -> count
        for r, c, _ in size_cells:
            row_counts[r] = row_counts.get(r, 0) + 1
            col_counts[c] = col_counts.get(c, 0) + 1

        max_row_count = max(row_counts.values()) if row_counts else 0
        max_col_count = max(col_counts.values()) if col_counts else 0

        # 가로형: 한 행에 사이즈가 2개 이상 모여있음
        # 세로형: 한 열에 사이즈가 2개 이상 모여있음
        if max_row_count >= 2 and max_row_count >= max_col_count:
            detected_format = "horizontal"
            # 가장 많은 사이즈가 있는 행을 사이즈 헤더 행으로 사용
            header_row = max(row_counts, key=row_counts.get)
            # 해당 행의 사이즈 셀들만 추출
            header_sizes = [(r, c, s) for r, c, s in size_cells if r == header_row]

            # 수량: 사이즈 셀 바로 아래 행에서 추출
            sizes_result = {}
            for r, c, size in header_sizes:
                qty = 0
                # 아래 행에서 수량 찾기 (최대 3행 아래까지)
                for offset in range(1, 4):
                    next_r = r + offset
                    if next_r < len(rows_data) and c < len(rows_data[next_r]):
                        qty = _extract_quantity(rows_data[next_r][c])
                        if qty > 0:
                            break
                # 중복 사이즈 수량 합산
                if size in sizes_result:
                    sizes_result[size] += qty
                else:
                    sizes_result[size] = qty

        elif max_col_count >= 2:
            detected_format = "vertical"
            # 가장 많은 사이즈가 있는 열을 사이즈 열로 사용
            size_col = max(col_counts, key=col_counts.get)
            col_sizes = [(r, c, s) for r, c, s in size_cells if c == size_col]

            # 수량: 사이즈 셀 바로 오른쪽 열에서 추출
            sizes_result = {}
            for r, c, size in col_sizes:
                qty = 0
                # 오른쪽 열에서 수량 찾기 (최대 3열 오른쪽까지)
                for offset in range(1, 4):
                    next_c = c + offset
                    if r < len(rows_data) and next_c < len(rows_data[r]):
                        qty = _extract_quantity(rows_data[r][next_c])
                        if qty > 0:
                            break
                if size in sizes_result:
                    sizes_result[size] += qty
                else:
                    sizes_result[size] = qty

        else:
            # 사이즈가 1개만 있거나 분산되어 있는 경우
            detected_format = "unknown"
            sizes_result = {}
            for r, c, size in size_cells:
                # 오른쪽과 아래 모두 탐색
                qty = 0
                # 오른쪽 먼저
                if r < len(rows_data) and c + 1 < len(rows_data[r]):
                    qty = _extract_quantity(rows_data[r][c + 1])
                # 오른쪽에 없으면 아래
                if qty == 0 and r + 1 < len(rows_data) and c < len(rows_data[r + 1]):
                    qty = _extract_quantity(rows_data[r + 1][c])
                if size in sizes_result:
                    sizes_result[size] += qty
                else:
                    sizes_result[size] = qty

        # 결과 정리: SIZE_KEYWORDS 순서대로 정렬
        ordered_sizes = []
        for kw in SIZE_KEYWORDS:
            if kw in sizes_result:
                ordered_sizes.append({
                    "size": kw,
                    "quantity": sizes_result[kw],
                })

        total_qty = sum(s["quantity"] for s in ordered_sizes)

        result = {
            "success": True,
            "sizes": ordered_sizes,
            "total_quantity": total_qty,
            "source_sheet": sheet_name,
            "detected_format": detected_format,
        }

        # 가장 많은 사이즈를 찾은 시트를 최종 결과로 사용
        if best_result is None or len(ordered_sizes) > len(best_result["sizes"]):
            best_result = result

    wb.close()

    if best_result is None:
        return {
            "success": False,
            "error": "엑셀 파일에서 사이즈 정보를 찾을 수 없습니다. 셀에 S, M, L, XL 등의 사이즈 키워드가 있는지 확인하세요.",
            "sizes": [],
            "total_quantity": 0,
            "source_sheet": "",
            "detected_format": "unknown",
        }

    return best_result
